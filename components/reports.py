# =========================================================================
# AUTOMATED SURVEILLANCE & BIOASSAY REPORTING (components/reports.py)
#
# Previously crashed on load (KeyError on LGA_District, Collection_Date,
# etc. — none of which exist on specimen_records) and had a duplicate
# render_reports_page() definition that silently discarded the header
# call. Rebuilt entirely around the real schema: specimen_records for
# surveillance exports, bioassay_results for resistance exports.
# =========================================================================
import datetime
import io

import pandas as pd
import streamlit as st
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from utils.data_manager import (
    add_collector_column,
    as_screening_dict,
    batch_catch_summary,
    classify_resistance_status,
    compute_mortality_percentage,
    extract_genus_counts_from_screening,
    load_bioassay_results,
    load_specimen_records,
)
from utils.icons import render_page_header
from utils.pcr_and_accuracy import render_specimen_qr


# =========================================================================
# Helpers — specimen surveillance side
# =========================================================================
def _style_workbook(workbook) -> None:
    """Bold the header row and widen every column to fit its content.

    Column widths are keyed off the column's index rather than col[0].column_letter: the
    first cell of a column can be a MergedCell, which has no column_letter at all, so that
    lookup raises AttributeError on any sheet with a merged cell in row 1.
    """
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for idx, col in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(idx)].width = max(max_len + 3, 12)


def _flatten_specimen_df(df: pd.DataFrame) -> pd.DataFrame:
    """Adds readable genus count and identification-summary columns derived
    from field_screening_result, without mutating the caller's copy."""
    out = df.copy()
    if "collection_date" in out.columns:
        out["collection_date"] = pd.to_datetime(out["collection_date"], errors="coerce")
    else:
        out["collection_date"] = pd.NaT

    genus_totals = {"Anopheles": [], "Culex": [], "Aedes": [], "Other": []}
    methods = []
    for _, row in out.iterrows():
        counts = extract_genus_counts_from_screening(row.get("field_screening_result"))
        for g in genus_totals:
            genus_totals[g].append(counts.get(g, 0))
        # Decoded rather than isinstance-checked: field_screening_result is JSONB and does
        # not always arrive parsed, and a string yielded screening_method=None — a blank
        # column in the report and the export, with nothing to indicate it was a parse
        # miss rather than a record genuinely lacking a method.
        result = as_screening_dict(row.get("field_screening_result"))
        methods.append(result.get("screening_method") or None)

    for g, values in genus_totals.items():
        out[g] = values
    out["screening_method"] = methods
    return out


GENUS_COLS = ["Anopheles", "Culex", "Aedes", "Other"]

NO_LGA = "Not recorded"


def _lga_series(df: pd.DataFrame) -> pd.Series:
    """The lga column with every way of saying "nothing" collapsed into one label.

    A missing LGA arrives as NULL from some write paths and as an empty string from others.
    Left as they are, the blank counts as an LGA in "LGAs covered" and splits the per-LGA
    table into a "Not recorded" row and a nameless one — two rows, one absence.
    """
    lga = df["lga"] if "lga" in df.columns else pd.Series(index=df.index, dtype="object")
    return lga.astype("object").where(lga.notna(), NO_LGA).apply(
        lambda v: str(v).strip() or NO_LGA
    )


def _totals(df: pd.DataFrame) -> dict:
    """The headline numbers, counted as specimens rather than rows.

    A row is not a mosquito: one manual_field_log row is a whole collection event holding
    several hundred. Reporting row counts to a reader who is not going to ask which is which
    is how "171 records" gets read as 171 mosquitoes.
    """
    present = [c for c in GENUS_COLS if c in df.columns]
    counts = {c: int(df[c].sum()) for c in present}
    total = sum(counts.values())
    confirmed = int((df["pcr_status"] == "confirmed").sum()) if "pcr_status" in df.columns else 0
    return {
        "counts": counts,
        "total": total,
        "anopheles": counts.get("Anopheles", 0),
        "anopheles_share": (counts.get("Anopheles", 0) / total * 100) if total else 0.0,
        "confirmed": confirmed,
        "records": len(df),
        # Named LGAs only: "Not recorded" is the absence of one, and counting it inflated
        # "LGAs covered" by one for any selection with a blank.
        "lgas": sorted(v for v in _lga_series(df).unique() if v != NO_LGA),
    }


def _plain_summary(df: pd.DataFrame) -> str:
    """One paragraph, in words, for a reader who will not read the table.

    Every figure comes from the filtered data. Where there is nothing to say — no PCR
    confirmations, no LGA recorded — it says so plainly instead of omitting the sentence and
    leaving the impression the question was never asked.
    """
    t = _totals(df)
    if not t["total"]:
        return (
            "No mosquitoes were counted in the records matching these filters. This can mean "
            "no collections took place, or that the entries recorded no genus counts."
        )

    dates = df["collection_date"].dropna() if "collection_date" in df.columns else pd.Series(dtype="datetime64[ns]")
    when = ""
    if not dates.empty:
        first, last = dates.min().date(), dates.max().date()
        when = f" between {first:%d %B %Y} and {last:%d %B %Y}" if first != last else f" on {first:%d %B %Y}"

    where = ""
    if t["lgas"]:
        named = ", ".join(t["lgas"][:3])
        more = f" and {len(t['lgas']) - 3} other LGA(s)" if len(t["lgas"]) > 3 else ""
        where = f" in {named}{more}"
    else:
        where = " (no LGA recorded on these entries)"

    parts = [
        f"**{t['total']:,} mosquitoes** were collected{when}{where}, "
        f"across **{t['records']:,}** recorded entries."
    ]
    parts.append(
        f"**{t['anopheles']:,} ({t['anopheles_share']:.0f}%)** were *Anopheles*, the genus that "
        "transmits malaria — this is the figure that matters for malaria risk."
    )
    if t["confirmed"]:
        parts.append(
            f"**{t['confirmed']:,}** specimen(s) have been confirmed by PCR, the laboratory test "
            "that identifies species which look identical under a microscope."
        )
    else:
        parts.append(
            "No specimens have been PCR-confirmed yet, so species-level identifications here "
            "remain provisional."
        )
    return " ".join(parts)


def _lga_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Counts per LGA — a handful of meaningful rows, instead of the first 20 of hundreds."""
    if "lga" not in df.columns:
        return pd.DataFrame()
    present = [c for c in GENUS_COLS if c in df.columns]
    if not present:
        return pd.DataFrame()
    out = df.groupby(_lga_series(df))[present].sum().reset_index()
    out.columns = ["LGA"] + present
    out["Total"] = out[present].sum(axis=1)
    return out.sort_values("Total", ascending=False).reset_index(drop=True)


def _first_text(*values, default: str) -> str:
    """The first value that is real, readable text — skipping None, NaN and blanks."""
    for value in values:
        if value is None or not pd.notna(value):
            continue
        text = str(value).strip()
        if text:
            return text
    return default


def _first_photo_url(photo_urls) -> str | None:
    if isinstance(photo_urls, list) and len(photo_urls) > 0:
        return photo_urls[0]
    return None


def _entries_with_photos(df: pd.DataFrame) -> pd.DataFrame:
    """Just the entries carrying a photo, with the first one resolved into `_first_photo`.

    The photo-evidence section shows one entry at a time, so an entry with no photo has
    nothing to contribute there — it only padded the picker with rows that opened onto
    "No photo for this entry". Every entry, photo or not, is still listed in the detailed
    table above.

    `_first_photo` is guaranteed to be a real non-empty URL string on every row returned:
    under pandas 3 the applied column takes the new `str` dtype and a photo-less row's None
    comes back out as float NaN — which is truthy, so a plain `if` passed it straight to
    st.image() and the page died with "'float' object has no attribute 'format'".
    """
    if "photo_urls" not in df.columns:
        return df.iloc[0:0].assign(_first_photo=pd.Series(dtype="object"))
    out = df.copy()
    out["_first_photo"] = [_first_photo_url(v) for v in out["photo_urls"]]
    keep = out["_first_photo"].apply(lambda v: isinstance(v, str) and bool(v))
    return out[keep]


def _entry_label(row) -> str:
    """The one-line summary identifying an entry in the photo-evidence picker.

    The date is checked with isinstance(pd.Timestamp), not hasattr("date"):
    _flatten_specimen_df coerces the column with errors="coerce", so a missing date arrives
    as pd.NaT — which *has* a .date() returning NaT, and is truthy besides. An attribute
    check and an `or` fallback both sailed past it and the row read "NaT · Jere · 4".
    """
    date = row.get("collection_date")
    date_text = date.date() if isinstance(date, pd.Timestamp) else "date n/a"
    # notna, not `or`: a None in an object column comes back as float NaN, which is truthy,
    # so plain `a or b or "fallback"` selects the NaN and the header reads "· nan ·" —
    # the same pandas trap as the NaT above and the photo URL below.
    place = _first_text(row.get("lga"), row.get("breeding_site_type"), default="location n/a")
    return f"{date_text} · {place} · {_catch_phrase(row)}"


def _catch_phrase(row) -> str:
    """How much this entry represents, said the way its own row means it.

    A batch is described by what it *caught*, not by what is left on it. The flattened genus
    columns hold the batch net of its vialed-out children — correct for totals, and actively
    misleading here: a fully vialed batch nets to zero, so the photo of a collection event
    that yielded 170 mosquitoes was captioned "0 specimen(s)". Where the catch has gone is
    said out loud instead, since a reader comparing this to the totals above deserves to know
    those 170 are counted as 170 individual rows and not lost.
    """
    batch = batch_catch_summary(row.get("field_screening_result"))
    if batch is None:
        caught = sum(int(row.get(g, 0) or 0) for g in GENUS_COLS)
        return f"{caught:,} specimen" + ("" if caught == 1 else "s")

    collected, vialed = batch["collected"], batch["vialed_out"]
    if not vialed:
        return f"{collected:,} collected"
    if batch["remaining"]:
        return f"{collected:,} collected ({vialed:,} vialed out)"
    return f"{collected:,} collected (all vialed out)"


def _genus_counts_text(row) -> str:
    """Per-genus counts for one entry, saying the same thing its label says.

    A batch is shown its raw catch for the same reason `_catch_phrase` reports one: the
    flattened genus columns hold the batch net of its vialed-out children, so a fully vialed
    batch labelled "170 collected" had its own detail pane read "Anopheles 0, Culex 0,
    Aedes 0" two lines beneath — the entry contradicting itself.
    """
    batch = batch_catch_summary(row.get("field_screening_result"))
    counts = batch["by_genus"] if batch else {g: int(row.get(g, 0) or 0) for g in GENUS_COLS}
    return ", ".join(f"{g} {counts.get(g, 0):,}" for g in ("Anopheles", "Culex", "Aedes"))


def _compile_specimen_excel(df: pd.DataFrame) -> io.BytesIO:
    buffer = io.BytesIO()

    total_records = len(df)
    genus_cols = [c for c in ["Anopheles", "Culex", "Aedes", "Other"] if c in df.columns]
    genus_totals = df[genus_cols].sum() if genus_cols else pd.Series(dtype=float)
    pcr_counts = df["pcr_status"].value_counts() if "pcr_status" in df.columns else pd.Series(dtype=int)

    # Records and specimens are different numbers, and the summary used to report the row
    # count under "Total Specimens Logged". One manual_field_log row is a whole batch — 500
    # mosquitoes on one row — so that headline could read 3 for a catch of 1,570 while the
    # per-genus totals beneath it, which are summed counts, said otherwise.
    summary_rows = [
        ("Total Specimens Caught", int(genus_totals.sum()) if genus_cols else 0),
        ("Records (collection events + identifications)", total_records),
    ]
    for g in genus_cols:
        summary_rows.append((f"Total {g} Count", int(genus_totals.get(g, 0))))
    for status, count in pcr_counts.items():
        summary_rows.append((f"PCR Status: {status}", int(count)))

    summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    site_breakdown = pd.DataFrame()
    if "breeding_site_type" in df.columns and genus_cols:
        site_breakdown = df.groupby(df["breeding_site_type"].fillna("Unspecified"))[genus_cols].sum().reset_index()
        site_breakdown.columns = ["Breeding Site Type"] + genus_cols

    # Per-LGA counts: the breakdown a health authority actually reports on, and the one that
    # lines up with what the DHIS2 export submits.
    lga_breakdown = pd.DataFrame()
    if "lga" in df.columns and genus_cols:
        lga_breakdown = df.groupby(df["lga"].fillna("Not recorded"))[genus_cols].sum().reset_index()
        lga_breakdown.columns = ["LGA"] + genus_cols

    # Resolve the collector name BEFORE dropping field_screening_result — the readable
    # label lives inside that JSON, so dropping it first would leave only a bare UUID in
    # the exported sheet.
    raw_export = add_collector_column(df).drop(columns=["field_screening_result"], errors="ignore").copy()
    if "photo_urls" in raw_export.columns:
        raw_export["photo_urls"] = raw_export["photo_urls"].apply(
            lambda x: ", ".join(x) if isinstance(x, list) else ""
        )

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Executive_Summary", index=False)
        if not lga_breakdown.empty:
            lga_breakdown.to_excel(writer, sheet_name="LGA_Breakdown", index=False)
        if not site_breakdown.empty:
            site_breakdown.to_excel(writer, sheet_name="Site_Type_Breakdown", index=False)
        raw_export.to_excel(writer, sheet_name="Raw_Records", index=False)

        _style_workbook(writer.book)

    buffer.seek(0)
    return buffer


# =========================================================================
# Helpers — bioassay / resistance side
# =========================================================================
def _flatten_bioassay_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["mortality_pct"] = out.apply(
        lambda r: compute_mortality_percentage(r.get("mortality_24hr", 0), r.get("mosquitoes_exposed", 0)),
        axis=1,
    )
    out["resistance_status"] = out["mortality_pct"].apply(classify_resistance_status)
    return out


def _compile_bioassay_excel(df: pd.DataFrame) -> io.BytesIO:
    buffer = io.BytesIO()

    summary = df.groupby(["treatment_name", "concentration_pct", "is_control"]).agg(
        total_exposed=("mosquitoes_exposed", "sum"),
        total_mortality=("mortality_24hr", "sum"),
        replicates=("replicate_number", "nunique"),
    ).reset_index()
    summary["mortality_pct"] = summary.apply(
        lambda r: compute_mortality_percentage(r["total_mortality"], r["total_exposed"]), axis=1
    )
    summary["resistance_status"] = summary["mortality_pct"].apply(classify_resistance_status)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Treatment_Summary", index=False)
        df.to_excel(writer, sheet_name="Raw_Replicates", index=False)

        _style_workbook(writer.book)

    buffer.seek(0)
    return buffer


# =========================================================================
# Page
# =========================================================================
def render_reports_page():
    render_page_header(
        title="Automated Reports",
        icon_name="reports",
        caption="Filter, review, and export specimen surveillance and bioassay resistance data.",
    )
    st.markdown("---")

    tab1, tab2 = st.tabs(["Specimen Surveillance Report", "Bioassay Resistance Report"])

    # ── TAB 1: specimen_records ──────────────────────────────────────────
    with tab1:
        raw_df = load_specimen_records()
        if raw_df.empty:
            st.warning("No specimen records yet. Log entries via Site Log Entry or Diagnostics first.")
        else:
            df = _flatten_specimen_df(raw_df)
            valid_dates = df.dropna(subset=["collection_date"])

            st.subheader("Filters")
            fcol1, fcol2, fcol3, fcol4 = st.columns(4)

            with fcol1:
                if not valid_dates.empty:
                    min_date, max_date = valid_dates["collection_date"].min().date(), valid_dates["collection_date"].max().date()
                    if min_date == max_date:
                        start_date, end_date = min_date, max_date
                        st.info(f"Single date in dataset: {min_date}")
                    else:
                        # st.date_input returns a 1-tuple between the two clicks of a range
                        # selection (and an empty one if cleared). Unpacking it straight
                        # into two names raises ValueError and takes the page down mid-pick.
                        picked = st.date_input(
                            "Date range", value=(min_date, max_date), min_value=min_date, max_value=max_date
                        )
                        if isinstance(picked, (tuple, list)) and len(picked) == 2:
                            start_date, end_date = picked
                        else:
                            # Mid-selection: keep the full window until they choose an end.
                            start_date, end_date = min_date, max_date
                else:
                    start_date, end_date = None, None
                    st.info("No valid collection dates in this dataset.")

            with fcol2:
                site_options = sorted(df["breeding_site_type"].dropna().unique()) if "breeding_site_type" in df.columns else []
                selected_sites = st.multiselect("Breeding site type (all if empty)", site_options)

            with fcol3:
                # LGA is the dimension health reporting is organised by — it is what the
                # DHIS2 export aggregates to — so it belongs alongside habitat here.
                lga_options = sorted(df["lga"].dropna().unique()) if "lga" in df.columns else []
                selected_lgas = st.multiselect("LGA (all if empty)", lga_options)

            with fcol4:
                pcr_options = sorted(df["pcr_status"].dropna().unique()) if "pcr_status" in df.columns else []
                selected_pcr = st.multiselect("PCR status (all if empty)", pcr_options)

            processed = df.copy()
            if start_date and end_date and "collection_date" in processed.columns:
                mask = processed["collection_date"].dt.date.between(start_date, end_date) | processed["collection_date"].isna()
                processed = processed[mask]
            if selected_sites:
                processed = processed[processed["breeding_site_type"].isin(selected_sites)]
            if selected_lgas:
                processed = processed[processed["lga"].isin(selected_lgas)]
            if selected_pcr:
                processed = processed[processed["pcr_status"].isin(selected_pcr)]

            st.markdown("---")
            st.subheader("Query Results")

            if processed.empty:
                st.error("No records match the current filters.")
            else:
                totals = _totals(processed)

                # In words first. The tables below answer "which sites, which dates"; this
                # answers "what happened", which is the only question most readers of this
                # page have.
                st.info(_plain_summary(processed))

                kpi1, kpi2, kpi3, kpi4 = st.columns(4)
                kpi1.metric(
                    "Mosquitoes collected", f"{totals['total']:,}",
                    help="Individual mosquitoes, not entries. One field log can hold hundreds.",
                )
                kpi2.metric(
                    "Anopheles (malaria vector)", f"{totals['anopheles']:,}",
                    delta=f"{totals['anopheles_share']:.0f}% of the catch", delta_color="off",
                )
                kpi3.metric(
                    "PCR confirmed", f"{totals['confirmed']:,}",
                    help="Specimens whose species was verified in the laboratory.",
                )
                kpi4.metric(
                    "LGAs covered", f"{len(totals['lgas']):,}",
                    help="Local Government Areas with at least one entry in this selection.",
                )

                # Per LGA, not the first 20 rows of hundreds. "Showing 20 of 171" invited the
                # reading that 171 was a number of mosquitoes, and 20 arbitrary rows answer no
                # question anyone has.
                by_lga = _lga_summary(processed)
                if not by_lga.empty:
                    st.markdown("**Where they were collected**")
                    st.dataframe(by_lga, width="stretch", hide_index=True)

                shown = add_collector_column(processed)
                display_cols = [c for c in [
                    "collection_date", "lga", "breeding_site_type", "screening_method", "Collector",
                    "Anopheles", "Culex", "Aedes", "Other", "pcr_status", "pcr_confirmed_species",
                ] if c in shown.columns]
                with st.expander(f"Every entry ({len(processed):,}) — the detailed table"):
                    st.caption(
                        "One row is one **entry**, not one mosquito: a field log row is a whole "
                        "collection event holding its counts, while a vialed-out individual is "
                        "a single specimen. The totals above already account for both."
                    )
                    st.dataframe(shown[display_cols], width="stretch", hide_index=True)

                st.markdown("---")
                st.subheader("Photo Evidence")

                evidence_df = _entries_with_photos(processed)
                if evidence_df.empty:
                    st.info("No entries with photos in the current filter.")
                else:
                    # A picker over the entries that actually have a photo, rather than a
                    # stack of expanders. Entries are looked at one at a time, so 25 of them
                    # filled several screens of the command centre to show one photo — and
                    # most were "No photo for this entry", which is not evidence. The
                    # detailed table above is where every entry is listed.
                    choice = st.selectbox(
                        f"Entry to view ({len(evidence_df):,} with a photo)",
                        options=range(len(evidence_df)),
                        # Indices, not label strings: two entries from the same site on the
                        # same day produce identical labels, and a selectbox keyed by value
                        # cannot tell them apart.
                        format_func=lambda i: _entry_label(evidence_df.iloc[i]),
                    )
                    row = evidence_df.iloc[choice]
                    photo_url = row["_first_photo"]

                    photo_col_, detail = st.columns([1, 1])
                    with photo_col_:
                        st.image(photo_url, width="stretch")
                    with detail:
                        st.write(f"**Site:** {_first_text(row.get('breeding_site_type'), default='Not recorded')}")
                        st.write(f"**LGA:** {_first_text(row.get('lga'), default='Not recorded')}")
                        st.write(f"**Counts** — {_genus_counts_text(row)}")
                        st.write(f"**PCR:** {row.get('pcr_status', 'not_submitted')}")
                        gps_lat, gps_lon = row.get("gps_lat"), row.get("gps_lon")
                        if pd.notna(gps_lat) and pd.notna(gps_lon):
                            st.caption(f"GPS {gps_lat}, {gps_lon}")
                        specimen_id = row.get("specimen_id")
                        if specimen_id and specimen_id != "n/a":
                            st.caption(f"Specimen `{specimen_id}`")
                            render_specimen_qr(specimen_id, key=f"qr_report_{specimen_id}", width=120)

                st.markdown("---")
                st.subheader("Export")
                ex1, ex2 = st.columns(2)
                timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")

                with ex1:
                    excel_buf = _compile_specimen_excel(processed)
                    st.download_button(
                        "Download Excel (.xlsx)", data=excel_buf,
                        file_name=f"specimen_report_{timestamp_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary", width="stretch",
                    )
                with ex2:
                    csv_export = add_collector_column(processed).drop(
                        columns=["field_screening_result"], errors="ignore"
                    )
                    st.download_button(
                        "Download CSV", data=csv_export.to_csv(index=False),
                        file_name=f"specimen_report_{timestamp_str}.csv",
                        mime="text/csv", width="stretch",
                    )

    # ── TAB 2: bioassay_results ───────────────────────────────────────────
    with tab2:
        bio_df = load_bioassay_results()
        if bio_df.empty:
            st.info("No bioassay results submitted yet. Use Bioassay Result Entry to log results.")
        else:
            flat = _flatten_bioassay_df(bio_df)

            st.subheader("Treatment Summary")
            summary = flat.groupby(["treatment_name", "concentration_pct", "is_control"]).agg(
                total_exposed=("mosquitoes_exposed", "sum"),
                total_mortality=("mortality_24hr", "sum"),
                replicates=("replicate_number", "nunique"),
            ).reset_index()
            summary["mortality_pct"] = summary.apply(
                lambda r: compute_mortality_percentage(r["total_mortality"], r["total_exposed"]), axis=1
            )
            summary["resistance_status"] = summary["mortality_pct"].apply(classify_resistance_status)

            st.dataframe(
                summary.rename(columns={
                    "treatment_name": "Treatment", "concentration_pct": "Concentration (%)",
                    "is_control": "Control", "total_exposed": "Total Exposed",
                    "total_mortality": "Total Mortality", "replicates": "Replicates",
                    "mortality_pct": "Mortality (%)", "resistance_status": "Status",
                }),
                width="stretch", hide_index=True,
            )
            st.caption(
                "Mortality shown is raw, not Abbott's-corrected. Where control mortality "
                "exceeds 5%, results should be corrected against the matching control "
                "before drawing resistance conclusions (WHO 2016 guidelines)."
            )

            st.markdown("---")
            st.subheader("Export")
            timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")
            bio_excel = _compile_bioassay_excel(flat)
            st.download_button(
                "Download Bioassay Report (.xlsx)", data=bio_excel,
                file_name=f"bioassay_report_{timestamp_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", width="stretch",
            )
